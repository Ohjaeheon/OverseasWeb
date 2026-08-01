import os
import re
import io
import openpyxl
import msoffcrypto
from msoffcrypto.format.ooxml import OOXMLFile

# 설정
PASSWORD = "gotjsqn"
TEMPLATE_NAME = "양식.xlsx"

def get_week_info(folder_name):
    """폴더명에서 주차 정보(예: 6월4주)를 추출하고, 없으면 폴더명 자체를 반환합니다."""
    match = re.search(r"(\d+월\s*\d+주)", folder_name)
    if match:
        return match.group(1).replace(" ", "")
    return folder_name

def find_target_sheet(filename, sheetnames):
    """
    지역별 엑셀 파일명(예: '01.도쿄_주간예배...')의 번호 접두사를 분석하여
    템플릿의 시트명(예: '1. 도쿄')과 매칭합니다.
    """
    basename = os.path.basename(filename)
    m = re.match(r"^(\d+)", basename)
    if not m:
        return None
    file_num = int(m.group(1))
    
    for sname in sheetnames:
        sm = re.match(r"^(\d+)", sname)
        if sm and int(sm.group(1)) == file_num:
            return sname
    return None

def load_encrypted_workbook(path, password):
    decrypted = io.BytesIO()
    with open(path, "rb") as f:
        file = msoffcrypto.OfficeFile(f)
        file.load_key(password=password)
        file.decrypt(decrypted)
    decrypted.seek(0)
    return openpyxl.load_workbook(decrypted, data_only=True, rich_text=True)

def copy_sheet_data_openpyxl(ws_src, ws_dst):
    extracted_date = None
    try:
        # A1 셀 복사
        a1_val = ws_src["A1"].value
        ws_dst["A1"].value = a1_val
        if a1_val:
            m = re.search(r"(\([^)]+\))", str(a1_val))
            if m:
                extracted_date = m.group(1)

        # C6:W13 범위의 데이터 복사 (수식 열 G, H, O, P, T, U 제외)
        ranges = [
            (3, 6),
            (9, 11),
            (12, 14),
            (17, 19),
            (22, 23)
        ]
        
        for r in range(6, 14):
            for start_col, end_col in ranges:
                for c in range(start_col, end_col + 1):
                    val = ws_src.cell(row=r, column=c).value
                    ws_dst.cell(row=r, column=c).value = val
    except Exception as e:
        print(f"    [오류] 데이터 복사 중 실패: {e}")
    return extracted_date

def update_common_sheets_date(wb, new_date):
    """공통 시트(해외-전체예배출석현황, 해외-자장부청)의 A1 셀 날짜 부분을 업데이트합니다."""
    for sname in ["해외-전체예배출석현황", "해외-자장부청"]:
        try:
            if sname in wb.sheetnames:
                sh = wb[sname]
                a1_val = sh["A1"].value
                if a1_val:
                    updated_val = re.sub(r"\([^)]+\)", new_date, str(a1_val))
                    sh["A1"].value = updated_val
                    print(f"    [공통] '{sname}' A1 날짜 업데이트 완료: {new_date}")
        except Exception as e:
            print(f"    [경고] '{sname}' A1 날짜 업데이트 중 오류: {e}")

def process_merge():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(base_dir, TEMPLATE_NAME)
    
    if not os.path.exists(template_path):
        print(f"[오류] 메인 경로에 '{TEMPLATE_NAME}' 파일이 존재하지 않습니다.")
        return

    # 하위 폴더 탐색
    subfolders = []
    for entry in os.scandir(base_dir):
        if entry.is_dir() and not entry.name.startswith('.'):
            has_regional = False
            for f in os.scandir(entry.path):
                if f.is_file() and f.name.endswith('.xlsx'):
                    if re.match(r'^\d+\.', f.name):
                        has_regional = True
                        break
            if has_regional:
                subfolders.append(entry)
                
    if not subfolders:
        print("[정보] 작업 대상이 되는 하위 폴더를 찾지 못했습니다.")
        return
        
    print(f"[정보] 총 {len(subfolders)}개의 작업 대상 폴더 발견: {[f.name for f in subfolders]}")
    
    try:
        # 템플릿 복호화
        print("[진행] 공통 엑셀 템플릿 파일 로드 중...")
        template_data = io.BytesIO()
        with open(template_path, "rb") as f:
            file = msoffcrypto.OfficeFile(f)
            file.load_key(password=PASSWORD)
            file.decrypt(template_data)
        
        for folder in subfolders:
            folder_name = folder.name
            week_info = get_week_info(folder_name)
            print(f"\n==================================================")
            print(f" 작업 시작: {folder_name} (추출된 주차: {week_info})")
            print(f"==================================================")
            
            final_주일_enc = os.path.join(folder.path, f"해외-예배출결현황({week_info})_주일.xlsx")
            final_수요_enc = os.path.join(folder.path, f"해외-예배출결현황({week_info})_수요.xlsx")
            
            # 결과 파일이 열려있는지 권한 체크 및 사전 비우기
            try:
                if os.path.exists(final_주일_enc):
                    os.remove(final_주일_enc)
                if os.path.exists(final_수요_enc):
                    os.remove(final_수요_enc)
            except PermissionError:
                print(f"[오류] 취합 결과 파일이 이미 열려 있어 접근할 수 없습니다.")
                print(f"작업 폴더 내의 엑셀 창을 완전히 닫고 다시 실행해 주세요.")
                return
            
            # Load template workbooks from decrypted stream (data_only=False to keep formulas)
            template_data.seek(0)
            wb_주일 = openpyxl.load_workbook(template_data, rich_text=True)
            
            template_data.seek(0)
            wb_수요 = openpyxl.load_workbook(template_data, rich_text=True)
            
            # 하위 폴더 내 19개 지역별 파일 목록
            regional_files = [f.name for f in os.scandir(folder.path) 
                              if f.is_file() and f.name.endswith('.xlsx') and re.match(r'^\d+\.', f.name)]
            regional_files.sort()
            
            print(f"[정보] '{folder_name}' 폴더 내 지역별 파일 개수: {len(regional_files)}개")
            
            dest_sheetnames = wb_주일.sheetnames
            
            extracted_date_주일 = None
            extracted_date_수요 = None
            
            for r_file in regional_files:
                r_path = os.path.join(folder.path, r_file)
                
                try:
                    # 지역별 파일 열기 (data_only=True to copy values)
                    wb_src = load_encrypted_workbook(r_path, PASSWORD)
                    
                    # 템플릿 시트 매칭
                    dst_sheet_name = find_target_sheet(r_file, dest_sheetnames)
                    if not dst_sheet_name:
                        print(f"  [경고] '{r_file}' 매칭되는 템플릿 시트를 찾을 수 없어 건너뜁니다.")
                        continue
                    
                    # 주일 취합
                    has_주일 = False
                    if "(주일)예배출석현황" in wb_src.sheetnames:
                        has_주일 = True
                        sh_src = wb_src["(주일)예배출석현황"]
                        sh_dst = wb_주일[dst_sheet_name]
                        date_found = copy_sheet_data_openpyxl(sh_src, sh_dst)
                        if date_found:
                            if not extracted_date_주일:
                                extracted_date_주일 = date_found
                            elif date_found != extracted_date_주일:
                                print(f"  [경고] '{r_file}'의 주일 날짜({date_found})가 기준 날짜({extracted_date_주일})와 다릅니다.")
                        print(f"  [완료] '{r_file}' -> '{dst_sheet_name}' (주일) 값 병합 완료")
                    if not has_주일:
                        print(f"  [정보] '{r_file}' 파일에 '(주일)예배출석현황' 시트가 없어 해당 국가(시트)는 건너뜀")
                        
                    # 수요 취합
                    has_수요 = False
                    if "(수요)예배출석현황" in wb_src.sheetnames:
                        has_수요 = True
                        sh_src = wb_src["(수요)예배출석현황"]
                        sh_dst = wb_수요[dst_sheet_name]
                        date_found = copy_sheet_data_openpyxl(sh_src, sh_dst)
                        if date_found:
                            if not extracted_date_수요:
                                extracted_date_수요 = date_found
                            elif date_found != extracted_date_수요:
                                print(f"  [경고] '{r_file}'의 수요 날짜({date_found})가 기준 날짜({extracted_date_수요})와 다릅니다.")
                        print(f"  [완료] '{r_file}' -> '{dst_sheet_name}' (수요) 값 병합 완료")
                    if not has_수요:
                        print(f"  [정보] '{r_file}' 파일에 '(수요)예배출석현황' 시트가 없어 해당 국가(시트)는 건너뜜")
                        
                except Exception as e:
                    print(f"  [오류] '{r_file}' 파일 처리 중 에러 발생: {e}")
            
            # 공통 시트 날짜 업데이트
            if extracted_date_주일:
                update_common_sheets_date(wb_주일, extracted_date_주일)
            if extracted_date_수요:
                update_common_sheets_date(wb_수요, extracted_date_수요)
            
            # Save and encrypt 주일 file
            temp_out = io.BytesIO()
            wb_주일.save(temp_out)
            temp_out.seek(0)
            with open(final_주일_enc, "wb") as f_out:
                file = OOXMLFile(temp_out)
                file.encrypt(PASSWORD, f_out)
                
            # Save and encrypt 수요 file
            temp_out = io.BytesIO()
            wb_수요.save(temp_out)
            temp_out.seek(0)
            with open(final_수요_enc, "wb") as f_out:
                file = OOXMLFile(temp_out)
                file.encrypt(PASSWORD, f_out)

            print(f"  [성공] 생성 완료: {os.path.basename(final_주일_enc)}")
            print(f"  [성공] 생성 완료: {os.path.basename(final_수요_enc)}")
            
    except Exception as e:
        print(f"[오류] 대형 에러 발생: {e}")
        
    print("\n[알림] 모든 취합 작업이 성료되었습니다!")

if __name__ == "__main__":
    process_merge()
