import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
wb = openpyxl.load_workbook('모델링/데이터.xlsx')
print('시트 목록:', wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f'\n=== {name} 시트 ===')
    print(f'최대 행: {ws.max_row}, 최대 열: {ws.max_column}')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        print(row)
